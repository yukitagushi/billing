from __future__ import annotations

import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import range_boundaries

from .normalizer import to_excel_value


@dataclass
class WriteResult:
    output_files: List[Path]
    mapping_notes: List[Dict[str, str]]


def load_mapping(mapping_path: Path) -> Dict[str, Any]:
    if not mapping_path.exists():
        raise FileNotFoundError(f"mapping file not found: {mapping_path}")
    with mapping_path.open("r", encoding="utf-8") as f:
        data = yaml.safe_load(f) or {}
    if "templates" not in data:
        data["templates"] = []
    return data


def infer_value_type(fmt: str) -> str:
    lowered = (fmt or "").lower()
    if "和暦" in fmt:
        return "date_wareki"
    if "yyyy" in lowered or "日付" in fmt:
        return "date_iso"
    if "金額" in fmt:
        return "currency"
    if "数字" in fmt:
        return "number"
    if "checkbox" in lowered or "チェック" in fmt:
        return "checkbox"
    if "改行" in fmt or "複数行" in fmt:
        return "text_multiline"
    return "text"


def _parse_target_segment(segment: str, default_sheet: Optional[str] = None) -> Optional[Tuple[str, str]]:
    cleaned = (segment or "").strip()
    if not cleaned:
        return None
    if "!" in cleaned:
        sheet_name, cell_ref = cleaned.split("!", 1)
        return sheet_name.strip(), cell_ref.strip()
    if default_sheet:
        return default_sheet, cleaned
    return None


def parse_targets(cell_spec: str, default_sheet: Optional[str] = None) -> List[Tuple[str, str]]:
    if not cell_spec:
        return []
    segments = [seg.strip() for seg in cell_spec.replace("\n", ";").split(";")]
    targets: List[Tuple[str, str]] = []
    for segment in segments:
        parsed = _parse_target_segment(segment, default_sheet=default_sheet)
        if parsed:
            targets.append(parsed)
    return targets


def _fill_linear_range(ws: Worksheet, cell_range: str, value: Any) -> None:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    if min_col == max_col and min_row == max_row:
        ws.cell(row=min_row, column=min_col, value=value)
        return

    text = "" if value is None else str(value)
    chars = list(text)
    idx = 0

    if min_row == max_row:
        for col in range(min_col, max_col + 1):
            cell_value = chars[idx] if idx < len(chars) else ""
            ws.cell(row=min_row, column=col, value=cell_value)
            idx += 1
        return

    if min_col == max_col:
        for row in range(min_row, max_row + 1):
            cell_value = chars[idx] if idx < len(chars) else ""
            ws.cell(row=row, column=min_col, value=cell_value)
            idx += 1
        return

    # 2D range fallback: fill row-major with characters.
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell_value = chars[idx] if idx < len(chars) else ""
            ws.cell(row=row, column=col, value=cell_value)
            idx += 1


def _write_to_target(ws: Worksheet, target: str, value: Any) -> None:
    if ":" in target:
        _fill_linear_range(ws, target, value)
    else:
        ws[target] = value


def write_templates(
    *,
    answers_norm: Dict[str, str],
    schema_lookup: Dict[str, Dict[str, Any]],
    template_dir: Path,
    output_dir: Path,
    mapping_path: Path,
) -> WriteResult:
    mapping_data = load_mapping(mapping_path)
    output_dir.mkdir(parents=True, exist_ok=True)

    output_files: List[Path] = []
    notes: List[Dict[str, str]] = []

    for template in mapping_data.get("templates", []):
        source_file = str(template.get("source_file", "")).strip()
        output_file = str(template.get("output_file", "")).strip() or f"filled_{source_file}"
        template_key = str(template.get("template_key", source_file)).strip()

        source_path = template_dir / source_file
        if not source_path.exists():
            notes.append(
                {
                    "field_id": "*",
                    "template": template_key,
                    "level": "error",
                    "message": f"テンプレートファイルが見つかりません: {source_path.name}",
                }
            )
            continue

        if source_path.suffix.lower() == ".xls":
            notes.append(
                {
                    "field_id": "*",
                    "template": template_key,
                    "level": "warning",
                    "message": ".xls は直接編集できないため未出力です。convert_xls.sh で .xlsx 化してください。",
                }
            )
            continue

        workbook = load_workbook(source_path)

        explicit_mappings: Dict[str, Dict[str, Any]] = template.get("mappings", {}) or {}
        explicit_field_ids = set(explicit_mappings.keys())

        source_form_files: Sequence[str] = template.get("source_form_files") or [source_file]
        source_form_files = [str(x).strip() for x in source_form_files if str(x).strip()]

        # 1) explicit mappings
        for field_id, conf in explicit_mappings.items():
            answer = answers_norm.get(field_id, "")
            if answer == "":
                continue

            sheet_name = str(conf.get("sheet", "")).strip()
            cell_spec = str(conf.get("cell", "")).strip()
            value_type = str(conf.get("type", "text")).strip()

            targets = parse_targets(f"{sheet_name}!{cell_spec}" if sheet_name and "!" not in cell_spec else cell_spec)
            if not targets:
                notes.append(
                    {
                        "field_id": field_id,
                        "template": template_key,
                        "level": "warning",
                        "message": "マッピング先セルが不正です。",
                    }
                )
                continue

            excel_value = to_excel_value(answer, value_type)
            for target_sheet, target_cell in targets:
                if target_sheet not in workbook.sheetnames:
                    notes.append(
                        {
                            "field_id": field_id,
                            "template": template_key,
                            "level": "warning",
                            "message": f"シートが見つかりません: {target_sheet}",
                        }
                    )
                    continue
                ws = workbook[target_sheet]
                _write_to_target(ws, target_cell, excel_value)

        # 2) auto mapping from schema cell_range for remaining fields
        for field_id, answer in answers_norm.items():
            if answer == "" or field_id in explicit_field_ids:
                continue

            field = schema_lookup.get(field_id)
            if not field:
                continue

            form_file = str(field.get("form_file", "")).strip()
            if form_file not in source_form_files:
                continue

            cell_range = str(field.get("cell_range", "")).strip()
            if not cell_range:
                continue

            value_type = infer_value_type(str(field.get("format", "")))
            excel_value = to_excel_value(answer, value_type)

            targets = parse_targets(cell_range)
            if not targets:
                notes.append(
                    {
                        "field_id": field_id,
                        "template": template_key,
                        "level": "warning",
                        "message": f"cell_range の解析に失敗: {cell_range}",
                    }
                )
                continue

            for target_sheet, target_cell in targets:
                if target_sheet not in workbook.sheetnames:
                    notes.append(
                        {
                            "field_id": field_id,
                            "template": template_key,
                            "level": "warning",
                            "message": f"シートが見つかりません: {target_sheet}",
                        }
                    )
                    continue
                ws = workbook[target_sheet]
                _write_to_target(ws, target_cell, excel_value)

        out_path = output_dir / output_file
        workbook.save(out_path)
        output_files.append(out_path)

    return WriteResult(output_files=output_files, mapping_notes=notes)


def create_review_report(
    *,
    out_path: Path,
    validation_issues: List[Dict[str, Any]],
    mapping_notes: List[Dict[str, str]],
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "review"

    ws.append(["kind", "severity", "field_id", "message", "template"])

    for issue in validation_issues:
        ws.append(
            [
                "validation",
                issue.get("severity", "warning"),
                issue.get("field_id", ""),
                issue.get("message", ""),
                "",
            ]
        )

    for note in mapping_notes:
        ws.append(
            [
                "mapping",
                note.get("level", "warning"),
                note.get("field_id", ""),
                note.get("message", ""),
                note.get("template", ""),
            ]
        )

    wb.save(out_path)
    return out_path


def ensure_file(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(str(path))


def copy_if_missing(src: Path, dst: Path) -> None:
    if not dst.exists():
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(src, dst)
